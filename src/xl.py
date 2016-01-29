"""Put data into an excel workbook. Currently unused.
"""

from __future__ import division
from __future__ import print_function
from __future__ import unicode_literals

from openpyxl import load_workbook, Workbook

from ImsrgDataMap import ImsrgDataMap, Exp


def print_single_particle_energy_data_to_excel(e, hw, datadir, savepath,
                                               startrow=2):
    all_data_map = ImsrgDataMap(parent_directory=datadir)
    data_maps = all_data_map.map[Exp(e, hw)]
    index_orbital_map = data_maps.index_orbital_map
    ime_map = data_maps.index_mass_energy_map()

    try:
        wb = load_workbook(savepath)
    except IOError:
        wb = Workbook()

    ws = wb.active
    ws.title = 'e={e} hw={hw}'.format(e=e, hw=hw)

    row = startrow
    col = 1

    ws.cell(row=row, column=col).value = 'KEY'
    row += 1

    for i, s in zip(range(5), ['Index', 'n', 'l', 'j', 'tz']):
        ws.cell(row=row, column=col + i).value = s
    row += 1

    for oindex in sorted(index_orbital_map.keys()):
        ws.cell(row=row, column=col).value = int(oindex)
        qnums = index_orbital_map[oindex]
        for i, qn in zip(range(1, 5), qnums):
            ws.cell(row=row, column=col + i).value = qn
        row += 1
    row += 1

    ws.cell(row=row, column=col).value = 'DATA'
    row += 1

    ws.cell(row=row, column=col).value = 'Index'
    ws.cell(row=row, column=col + 1).value = 'A'
    ws.cell(row=row, column=col + 2).value = 'energy (MeV)'
    row += 1

    for oindex in sorted(ime_map.keys()):
        me_map = ime_map[oindex]

        for m in me_map.keys():
            ws.cell(row=row, column=col).value = int(oindex)
            ws.cell(row=row, column=col + 1).value = int(m)
            ws.cell(row=row, column=col + 2).value = me_map[m]
            row += 1

    wb.save(savepath)
