from __future__ import division
from __future__ import print_function

from matplotlib import pyplot as plt
from openpyxl import Workbook, load_workbook
from scipy.optimize import curve_fit as cf

from ImsrgDataMap import ImsrgDataMap, Exp
from constants import *
from fittransforms import *

E = 12
HW = 20

SHOW_INDIVIDUAL_FITS = True
COMPARE_RAW_DATA = True
COMPARE_FITS = True
COMPARE_PER_NUC_DATA = True
COMPARE_DERIV = True
COMPARE_LOG_LOG = True
PRINT_FITTING_PARAMETERS = True


def _map_to_arrays(m):
    """Convert a map of dimensionality 2 into an x and y array"""
    length = len(m)
    x = np.empty(length)
    y = np.empty(length)
    for k, i in zip(sorted(m.keys()), range(length)):
        x[i] = k
        y[i] = m[k]
    return x, y


def single_particle_deriv_curvefit(fitfn, e=E, hw=HW, **kwargs):
    return _single_particle_curvefit(fitfn, e, hw,
                                     code='spd', transform=derivative,
                                     **kwargs)


def single_particle_log_log_curvefit(fitfn, e=E, hw=HW, **kwargs):
    return _single_particle_curvefit(fitfn, e, hw,
                                     code='spll', transform=log_log,
                                     **kwargs)


def single_particle_identity_curvefit(fitfn, e=E, hw=HW, **kwargs):
    return _single_particle_curvefit(fitfn, e, hw,
                                     code='spi', transform=identity,
                                     **kwargs)


def single_particle_per_nucleon_curvefit(fitfn, e=E, hw=HW, **kwargs):
    return _single_particle_curvefit(fitfn, e, hw,
                                     code='sppn', transform=per_nucleon,
                                     **kwargs)


def _single_particle_curvefit(fitfn, e=E, hw=HW,
                              sourcedir=FILES_DIR,
                              savedir=PLOTS_DIR,
                              show_fits=False,
                              show_data_compare=False,
                              legend_data_compare=True,
                              show_rel_data_compare=False,
                              legend_rel_data_compare=True,
                              transform=identity,
                              verbose=False,
                              code='',
                              xlabel='A',
                              ylabel='Energy (MeV)',
                              sortkey=lambda k: k):
    all_data_map = ImsrgDataMap(parent_directory=sourcedir)
    data_maps = all_data_map.map[Exp(e, hw)]
    io_map = data_maps.index_orbital_map
    ime_map = data_maps.index_mass_energy_map()

    compare_data_plots = list()
    compare_fit_plots = list()
    orbital_fit_map = dict()

    for index in sorted(ime_map.keys(), key=sortkey):
        qnums = io_map[index]
        me_map = ime_map[index]

        xdata, ydata = _map_to_arrays(me_map)
        x, y = transform(xdata, ydata)
        compare_data_plots.append((x, y, index))

        popt, pcov = cf(fitfn, x, y)
        orbital_fit_map[qnums] = popt

        if verbose:
            print(str(index) + ': ' + str(qnums))
            print(popt, end='\n\n')

        fitx = np.linspace(x[0], x[-1])
        fity = np.array(list(map(lambda xi: fitfn(xi, *popt), fitx)))
        compare_fit_plots.append((fitx, fity, index))

        if show_fits:
            f = plt.figure()
            ax = f.add_subplot(111)
            ax.plot(xdata, ydata, '-b')
            ax.plot(fitx, fity, '-r')
            plt.plot(xdata, ydata, '-b')
            plt.plot(fitx, fity, '-r')
            title = ('{c}:Single particle energy {tr} fit using {fn} for '
                     'orbital {o} with e={e} '
                     'hw={hw}').format(c=code,
                                       tr=transform.__name__,
                                       fn=fitfn.__name__,
                                       o=index,
                                       e=e, hw=hw)
            plt.title(title)
            plt.xlabel(xlabel)
            plt.ylabel(ylabel)
            plt.savefig(savedir + '/' + title + '.png')
            plt.show()

    # DATA COMPARISONS
    title_temp = ('{c}:Comparison of {rel}{sp1}single particle energy {tr} '
                  '{dof}{sp2}{us}{sp2}{fn} with e={e} hw={hw}'
                  '').format(c=code,
                             tr=transform.__name__,
                             e=e, hw=hw,
                             rel='{rel}',
                             sp1='{sp1}',
                             fn='{fn}',
                             sp2='{sp2}',
                             dof='{dof}',
                             us='{us}')

    if show_data_compare:
        fdat = plt.figure()
        axdat = fdat.add_subplot(111)
        for dat in compare_data_plots:
            x, y, label = dat
            axdat.plot(x, y, '-', label=label)
            # plt.plot(x, y, '-', label=label)
        title = title_temp.format(rel='', sp1='', fn='', sp2='', dof='data',
                                  us='')
        _do_plot(title, xlabel, ylabel,
                 saveloc=savedir + '/' + title + '.png',
                 showlegend=legend_data_compare)
        plt.show()

        ffit = plt.figure()
        axfit = ffit.add_subplot(111)
        for fit in compare_fit_plots:
            x, y, label = fit
            axfit.plot(x, y, '-', label=label)
            # plt.plot(x, y, '-', label=label)
        title = title_temp.format(rel='', sp1='', fn=fitfn.__name__,
                                  sp2=' ', dof='fit', us='using')
        _do_plot(title, xlabel, ylabel,
                 saveloc=savedir + '/' + title + '.png',
                 showlegend=legend_data_compare)
        plt.show()

    if show_rel_data_compare:
        frdat = plt.figure()
        axrdat = frdat.add_subplot(111)
        for rdat in compare_data_plots:
            x, y, label = rdat
            y = y - y[0]
            axrdat.plot(x, y, '-', label=label)
            # plt.plot(x, y, '-', label=label)
        title = title_temp.format(rel='relative', sp1=' ', fn='',
                                  sp2='', dof='data', us='')
        _do_plot(title, xlabel, ylabel,
                 saveloc=savedir + '/' + title + '.png',
                 showlegend=legend_rel_data_compare)
        plt.show()

        frfit = plt.figure()
        axrfit = frfit.add_subplot(111)
        for rfit in compare_fit_plots:
            x, y, label = rfit
            y = y - y[0]
            axrfit.plot(x, y, '-', label=label)
            # plt.plot(x, y, '-', label=label)
        title = title_temp.format(rel='relative', sp1=' ', fn=fitfn.__name__,
                                  sp2=' ', dof='fit', us='using')
        _do_plot(title, xlabel, ylabel,
                 saveloc=savedir + '/' + title + '.png',
                 showlegend=legend_rel_data_compare)
        plt.show()

    return orbital_fit_map


def _do_plot(title, xlabel, ylabel, saveloc, showlegend):
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    if showlegend:
        plt.legend()
    plt.savefig(saveloc)


def print_single_particle_energy_data_to_excel(e, hw, datadir, savepath,
                                               startrow=2):
    all_data_map = ImsrgDataMap(parent_directory=datadir)
    data_maps = all_data_map.map[Exp(e, hw)]
    index_orbital_map = data_maps.index_orbital_map
    ime_map = data_maps.index_mass_energy_map()

    try:
        wb = load_workbook(savepath)
    except IOError:
        wb = Workbook()

    ws = wb.active
    ws.title = 'e={e} hw={hw}'.format(e=e, hw=hw)

    row = startrow
    col = 1

    ws.cell(row=row, column=col).value = 'KEY'
    row += 1

    for i, s in zip(range(5), ['Index', 'n', 'l', 'j', 'tz']):
        ws.cell(row=row, column=col + i).value = s
    row += 1

    for oindex in sorted(index_orbital_map.keys()):
        ws.cell(row=row, column=col).value = int(oindex)
        qnums = index_orbital_map[oindex]
        for i, qn in zip(range(1, 5), qnums):
            ws.cell(row=row, column=col + i).value = qn
        row += 1
    row += 1

    ws.cell(row=row, column=col).value = 'DATA'
    row += 1

    ws.cell(row=row, column=col).value = 'Index'
    ws.cell(row=row, column=col + 1).value = 'A'
    ws.cell(row=row, column=col + 2).value = 'energy (MeV)'
    row += 1

    for oindex in sorted(ime_map.keys()):
        me_map = ime_map[oindex]

        for m in me_map.keys():
            ws.cell(row=row, column=col).value = int(oindex)
            ws.cell(row=row, column=col + 1).value = int(m)
            ws.cell(row=row, column=col + 2).value = me_map[m]
            row += 1

    wb.save(savepath)


'''
def single_particle_energy_curvefit(fitfn, e=E, hw=HW,
                                    sourcedir=FILES_DIR,
                                    savedir=PLOTS_DIR,
                                    show_fits=SHOW_INDIVIDUAL_FITS,
                                    show_data_compare=COMPARE_RAW_DATA,
                                    show_fit_compare=COMPARE_FITS,
                                    show_data_compare2=COMPARE_PER_NUC_DATA,
                                    show_deriv_compare=COMPARE_DERIV,
                                    show_log_log_compare=COMPARE_LOG_LOG,
                                    verbose=PRINT_FITTING_PARAMETERS):
    """Perform a curve fit for each of the single particle orbitals, for
    the files in the given directory with the given e and hw values


    :param fitfn: Function to use for curve-fitting. Must be of the form
    f(x, a1, a2, a3, ... aN)
    :param e: the energy level
    :param hw: the hw frequency
    :param sourcedir: the parent directory containing related files
    :param show_fits: whether to show the individual orbital fits
    :param show_data_compare: whether to show a comparison between the
    single particle orbital energy dependencies on mass
    :param show_fit_compare: whether to show a comparison between the
    single particle fitted energy dependencies on mass
    :param show_deriv_compare: whether to show a comparison between the
    single particle first derivatives
    :param show_log_log_compare: whether to show a comparison between the log-log
    plots of the data
    :param verbose: whether to output the fitted data
    :return: a map from the QuantumNumbers of each orbital to its fit
    parameters
    """
    all_data_map = ImsrgDataMap(parent_directory=FILES_DIR)
    data_maps = all_data_map.map[Exp(E, HW)]
    index_orbital_map = data_maps.index_orbital_map
    ime_map = data_maps.index_mass_energy_map()

    compare_data_plots = list()
    compare_deriv_plots = list()
    compare_fits_plots = list()
    orbital_fit_map = dict()
    for index in sorted(ime_map.keys(), key=lambda k: index_orbital_map[k].j):
        qnums = index_orbital_map[index]

        me_map = ime_map[index]
        xdata, ydata = _map_to_arrays(me_map)
        compare_data_plots.append((xdata, ydata, qnums))

        xderiv = np.array(xdata)[:-1]
        yderiv = np.empty(shape=len(xderiv))
        for i in range(len(xderiv)):
            yderiv[i] = ydata[i+1] - ydata[i]
        compare_deriv_plots.append((xderiv, yderiv, qnums))

        try:
            popt, pcov = cf(fitfn, xdata, ydata)
            orbital_fit_map[index_orbital_map[index]] = popt
        except RuntimeError:
            print('RuntimeError for orbital {o}'.format(o=qnums))
            continue

        if verbose:
            print(qnums)
            print(popt)
            print()

        curve_x = np.linspace(xdata[0], xdata[-1])
        curve_y = np.array(list(map(lambda x: fitfn(x, *popt), curve_x)))
        compare_fits_plots.append((curve_x, curve_y, qnums))

        if show_fits:
            f = plt.figure()
            ax = f.add_subplot(111)
            ax.plot(xdata, ydata, '-b')
            ax.plot(curve_x, curve_y, '-r')
            title = ('Single particle energy fit using {fn} for '
                     'orbital {o} '
                     'with e={e} hw={hw}').format(fn=fitfn.__name__,
                                                  o=index,
                                                  e=e, hw=hw)
            plt.title(title)
            plt.xlabel('A')
            plt.ylabel('Energy (MeV)')
            plt.savefig(savedir + '/' + title + '.png')
            plt.show()

    if show_data_compare:
        for dat in compare_data_plots:
            x, y, label = dat
            y = y - y[0]
            plt.plot(x, y, '-', label=tuple(label))
            title = ('Single particle energies: '
                     'Comparison of relative energy-mass dependence '
                     'with e={e} hw={hw}').format(e=e, hw=hw)
            plt.title(title)
            plt.xlabel('A')
            plt.ylabel('Relative energy (MeV)')
            plt.savefig(savedir + '/' + title + '.png')
        plt.show()

    if show_fit_compare:
        for fit in compare_fits_plots:
            x, y, label = fit
            y = y - y[0]
            plt.plot(x, y, '-', label=tuple(label))
            title = ('Single particle energies: '
                     'Comparison of fits using {fit} with '
                     'e={e} hw={hw}').format(fit=fitfn.__name__,
                                             e=e, hw=hw)
            plt.title(title)
            plt.xlabel('A')
            plt.ylabel('Relative fit energy (MeV)')
            plt.savefig(savedir + '/' + title + '.png')
        plt.show()

    if show_data_compare2:
        for dat in compare_data_plots:
            x, y, label = dat
            for i in range(len(y)):
                y[i] = y[i] / x[i]
            #y = y - y[0]
            plt.plot(x, y, label=tuple(label))
            title = ('Single particle energies: '
                     'Comparison of relative energy per nucleon '
                     'dependence on mass with '
                     'e={e} hw={hw}').format(e=e, hw=hw)
            plt.title(title)
            plt.xlabel('A')
            plt.ylabel('Energy per nucleon (MeV)')
            plt.savefig(savedir + '/' + title + '.png')
        plt.show()

    if show_log_log_compare:
        for dat in compare_data_plots:
            x, y, label = dat
            plt.plot(np.log(abs(x)), np.log(abs(y)), label=tuple(label))
            title = ('Single particle energies: '
                     'Comparison of log-logs with '
                     'e={e} hw={hw}').format(e=e, hw=hw)
            plt.title(title)
            plt.xlabel('log(A)')
            plt.ylabel('log(energy)')
            plt.savefig(savedir + '/' + title + '.png')
        plt.show()

    if show_deriv_compare:
        for der in compare_deriv_plots:
            x, y, label = der
            plt.plot(x, y, '-', label=tuple(label))
            title = ('Single particle energies: '
                     'Comparison of rate of change of mass-energy dependence '
                     'with e={e} hw={hw}').format(e=e, hw=hw)
            plt.title(title)
            plt.xlabel('A')
            plt.ylabel('dE/dA (MeV / nucleon)')
            plt.savefig(savedir + '/' + title + '.png')
        plt.show()

    return orbital_fit_map
'''